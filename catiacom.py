"""
Requirements:
    1) Open CATIAV5 and a product file in .CATProduct or .3dxml format.
    2) Run the code.
"""

import win32com.client
import numpy as np
import openpyxl
import os
import time

#Calculate the start time
start_time = time.time()

# Connect to CATIA V5
catia = win32com.client.Dispatch("CATIA.Application")

# Current Directory
current_directory = os.getcwd()

# Assing the necessary variables
active_document = catia.activedocument.name
print("Active Document is: ", active_document)

# Check if catia.ActiveDocument
if catia.ActiveDocument is not None:
    active_document = catia.ActiveDocument.name
    # Now you can work with 'active_document'
else:
    print("No active document in CATIA.")

# print(current_directory + "\\" + active_document)
document_path = current_directory + "\\" + active_document
print("Assembly path: " "r'" + document_path)
product_document = catia.Documents.Open(document_path)
product = product_document.Product
prod_name = product.Name
print("Product Name: ", prod_name)
excel_path = current_directory + "\\" + "weights.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Sheet1']
    start_cell = 'A2'
    start_row, start_col = openpyxl.utils.cell.coordinate_from_string(start_cell)

except:
    wb = openpyxl.Workbook()
    ws = wb.active
    print("'weights.xlsx' file is not in the same directory with .CATProduct file.\n"
          "Instead of a new 'weights.xlsx' file has been created.")
    start_cell = 'A2'
    start_row, start_col = openpyxl.utils.cell.coordinate_from_string(start_cell)

inertia_product = product.GetTechnologicalObject("Inertia")
mass_product = inertia_product.Mass
print(f"Total virtual mass of {prod_name} is {mass_product}.It might not indicates the true weight. \n "
      f"Check the assigned material to the parts.")

# Feature Vectors
weight_matrix = []
name_matrix = []

# Function to recursively analyze subparts and access their properties
def analyze_subparts(component):
    global weight_matrix
    global name_matrix

    for child in component.Products:

        # Access properties (e.g., name, weight) for each subpart
        name = child.Name
        weight = None
        inertia_child = child.GetTechnologicalObject("Inertia")

        # Attempt to get the "Mass" property if it exists
        try:
            weight = inertia_child.Mass
            name_matrix = np.append(name, name_matrix)
            weight_matrix = np.append(weight, weight_matrix)
            print(len(name_matrix))
        except Exception as e:
            print(f"Error while retrieving weight for '{name}': {e}")
            np.append("None", weight_matrix)

        # Check if the component has sub-components
        if hasattr(child, "Products"):
            analyze_subparts(child)  # Recursively analyze sub-subparts

    for idx, value in enumerate(name_matrix, start=1):
        cell = ws.cell(row=idx + 1, column=1)
        cell.value = value
    for idx, value in enumerate(weight_matrix, start=1):
        cell = ws.cell(row=idx + 1, column=2)
        cell.value = value
    wb.save(excel_path)

# Start the analysis
analyze_subparts(product)

# Calculate end time and total duration til end of the code
end_time = time.time()
pass_time = end_time - start_time
pass_time = int(pass_time)

# End Message
print(f"Process has been finished. \n"
      f"Now open {excel_path} and copy A and B columns than paste it to the 'weights_2.xlsx' file.")
if pass_time >= 60:
    print(f"Duration was: {pass_time/60} minutes.")
else:
    print(f"Duration was: {pass_time} seconds.")
input("Press Enter to close...")
